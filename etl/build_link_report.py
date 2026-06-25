# -*- coding: utf-8 -*-
"""웹앱 링크 검증 엑셀 생성
열: 반 / 학번 / 이름 / 웹앱목록 링크 / 최신 링크(PDF 임베드) / 변화 여부 / PDF 표시URL / 텍스트≠하이퍼링크
출력: 연구대회\웹앱 링크 검증.xlsx
"""
import warnings; warnings.filterwarnings("ignore")
import os, json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RC = r"C:\Users\user\Downloads\연구대회"
APPS_X = os.path.join(RC, "웹앱 목록.xlsx")
ROSTER_X = os.path.join(RC, "학생 명단.xlsx")
ENRICH_J = r"D:\project_clone\alb\etl\enrich.json"
OUT = os.path.join(RC, "웹앱 링크 검증.xlsx")

def nid(v):
    if v is None: return None
    s = re.sub(r"[^0-9]", "", str(v))
    return s if len(s) == 4 else None

def norm_url(u):
    return (u or "").strip().rstrip("/")

def lastseg(u):
    return (u or "").rstrip("/").split("/")[-1].lower()

def real_text_mismatch(emb, vis):
    """표시 텍스트 URL과 임베드 href 의 '공유 ID' 자체가 다른 경우만 True.
    (대소문자 차이, 표시 텍스트 truncation 은 제외)"""
    if not emb or not vis:
        return False
    a, b = lastseg(emb), lastseg(vis)
    if not a or not b or a == b:
        return False
    if a.startswith(b) or b.startswith(a):  # 표시 텍스트가 잘린 경우
        return False
    return True

# roster
wb = openpyxl.load_workbook(ROSTER_X, read_only=True, data_only=True); ws = wb.active
roster = {}
rows = list(ws.iter_rows(values_only=True)); hdr = rows[0]
ci = {h: i for i, h in enumerate(hdr)}
def col(*kw):
    for i, h in enumerate(hdr):
        if any(k in str(h or "") for k in kw): return i
    return None
c_cls, c_id, c_nm = col("반"), col("학번"), col("이름")
for r in rows[1:]:
    sid = nid(r[c_id])
    if sid: roster[sid] = {"cls": str(r[c_cls]).strip() if r[c_cls] is not None else "",
                           "name": str(r[c_nm]).strip() if r[c_nm] else ""}
wb.close()

# apps
wb = openpyxl.load_workbook(APPS_X, read_only=True, data_only=True); ws = wb.active
rows = list(ws.iter_rows(values_only=True)); hdr = rows[0]
def acol(*kw):
    for i, h in enumerate(hdr):
        if any(k in str(h or "") for k in kw): return i
    return None
a_id, a_url = acol("학번"), acol("appUrl", "URL", "url")
applist = {}
for r in rows[1:]:
    sid = nid(r[a_id]) if a_id is not None else None
    if sid: applist[sid] = norm_url(r[a_url]) if a_url is not None and r[a_url] else ""
wb.close()

# enrich
enrich = {}
if os.path.exists(ENRICH_J):
    d = json.load(open(ENRICH_J, encoding="utf-8"))
    for s in (d.get("students", d) if isinstance(d, dict) else d):
        sid = nid(s.get("hakbun") or s.get("id"))
        if sid: enrich[sid] = s

# build
out = openpyxl.Workbook(); sh = out.active; sh.title = "웹앱 링크 검증"
headers = ["반", "학번", "이름", "웹앱목록 링크", "최신 링크(PDF 임베드)", "변화 여부", "PDF 표시URL", "텍스트≠하이퍼링크"]
sh.append(headers)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, len(headers)+1):
    cell = sh.cell(1, c); cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4"); cell.alignment = Alignment(horizontal="center"); cell.border = border

FILL = {"변경됨": "FFF2CC", "목록없음(PDF만)": "E2EFDA", "PDF링크없음(목록만)": "FCE4D6", "둘다없음": "F2F2F2"}
cnt = {}
for sid in sorted(roster):
    r = roster[sid]
    listu = applist.get(sid, "")
    e = enrich.get(sid, {})
    emb = norm_url(e.get("embeddedLink"))
    vis = norm_url(e.get("visibleLink"))
    latest = emb or listu
    if not listu and not emb: status = "둘다없음"
    elif listu and not emb:    status = "PDF링크없음(목록만)"
    elif emb and not listu:    status = "목록없음(PDF만)"
    elif emb == listu:         status = "동일"
    else:                      status = "변경됨"
    cnt[status] = cnt.get(status, 0) + 1
    textmis = "Y" if real_text_mismatch(emb, vis) else ""
    row = [r["cls"], sid, r["name"], listu, latest, status, vis, textmis]
    sh.append(row)
    ri = sh.max_row
    for c in range(1, len(headers)+1):
        sh.cell(ri, c).border = border
    if status in FILL:
        for c in range(1, len(headers)+1):
            sh.cell(ri, c).fill = PatternFill("solid", fgColor=FILL[status])

widths = [6, 8, 10, 46, 46, 18, 40, 14]
for i, w in enumerate(widths, 1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
sh.freeze_panes = "A2"
sh.auto_filter.ref = "A1:H%d" % sh.max_row
out.save(OUT)
print("wrote", OUT)
print("status counts:", json.dumps(cnt, ensure_ascii=False))
